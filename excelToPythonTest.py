####COLLECTION OF FUNCTIONS FOR ANALYSING RTQUIC DATA FROM EXCEL####
####ALEX PEDEN, DECEMBER 2018#####

from openpyxl import load_workbook


class RTQuicSheet(object):
    

    SECONDS_PER_CYCLE = 949
    def __init__(self, excel_workbook_name, sheet_name):
        #clear previous workbook
        self.wb = None
        self.excel_workbook_name = excel_workbook_name
        self.sheet_name = sheet_name
        ##set workbook as file object
        try:
            self.wb = load_workbook(self.excel_workbook_name)
        except IOError:
            self.wb = None
            print("Could not open file"+ self.excel_workbook_name)
        ##get sheet as file object
        try:
            self.sheet = self.wb[self.sheet_name]
        except:
            self.sheet = None
            print("Could not find sheet")
        self.row_list = []
        self.row_max = None
        self.time_to_max = None ##eventually will be in seconds
        self.threshold = None 
        self.lag = None ##eventually will be in seconds


    """gets a work book object"""
    def get_wb(self):
        
        return wb
    
    """
    reads a selected row from a sheet file object and returns a list of what's in it.
    row_start refers the column number i.e. A = 1, B = 2 etc where you want to start
    reading the row (left to right).
    """
    def set_row_list(self, row_start, row):
        print("row number sent to set_row_list "+str(row))
        self.row_list = []
        for i in range(row_start, 1000):        
            try:
                val = self.sheet.cell(row, column=i).value
                assert(type(val) == int)
            except:   
                print("A problem occurred reading data from a row")
                self.row_list = None
            if val == None:
                break
            self.row_list = self.row_list + [val]
        print(self.row_list[0:20])


    """
    reads from a row as a list of ints and/or floats and obtains maximum value
    and time (in hours) when max value occurred
    """
    def set_row_max(self):
        for elem in self.row_list:
            assert(type(elem) == int or type(elem) == float)
            self.row_max = max(self.row_list)

    def get_row_max(self):
        return self.row_max

    def set_time_to_max(self):
        ## in seconds
        self.time_to_max = self.row_list.index(self.row_max)*(RTQuicSheet.SECONDS_PER_CYCLE) 

    def get_time_to_max(self):
        return self.time_to_max

    """
    reads from a row as a list of ints and/or doubles and obtains first of
    three consecutive values that are three times above the baseline along with
    the time to reach this value, or else returns  "no values found"
    """
    def setThreshold(self):
        assert(type(self.row_list[2]) == int or type(self.row_list[2]) == float)
        self.threshold = self.row_list[2]*3

    def setLag(self):
        self.setThreshold()
        self.lag = None
        for i in range(len(self.row_list)-2):
            if (self.row_list[i] > self.threshold and
                self.row_list[i+1] > self.threshold and
                self.row_list[i+2] > self.threshold):
                self.lag = i*(RTQuicSheet.SECONDS_PER_CYCLE)##in seconds
                return           
        print ("no lagtime found for row ")
        
    def getLag(self):
            return  self.lag ##in seconds

    "takes a time in seconds and converts to h:m format"
    def hours(self, func):
        
        def wrap():
            try:
                hours = str(func()//3600) 
                minutes = str(func()%3600//60)
                return hours +" hours: "+ minutes + " minutes"
            except:
                print("Could not convert to hours")
                return("N/A")
        return wrap

    def __str__(self):
        return "Workbook: "+ self.excel_workbook_name + " Sheet: " + self.sheet_name
        








