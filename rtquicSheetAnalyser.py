from openpyxl import Workbook
from os import listdir
from RTQuicSheet import *

class Error(Exception):
    pass
class FileNotSavedError(Error):
    pass

PATH_TO_FILES = "C:\\Users\\user\\OneDrive - University of Edinburgh\\excelMaxRow\\RTQuIC_for_analysis"
list_of_files = (listdir(PATH_TO_FILES))

"""Analyses data in an excel file row by row
Puts results in a destination excel file"""
class SheetAnalyser(object):
    def __init__(self, raw_data, raw_data_sheet):
        self.analysis_num = 0
        self.row_num = 2 #row in destination excel file
        self.raw_data = raw_data
        self.raw_data_sheet = raw_data_sheet
        self.result_list = []
        try:
            self.rtquic1 = RTQuicSheet(self.raw_data, self.raw_data_sheet)
        except:
            print("Could not generate RTQuicSheet object for analysing data "+ self.raw_data)
        try:
            self.wb = Workbook()
        except:
            print("Could not generate destination workbook object for " + self.raw_data)
        try:
            self.ws = self.wb.active
        except:
            print("Could not generate active worksheet within distination workbook for " + self.raw_data)        
    def set_analysis_num(self, num):
        self.analysis_num = num
    def set_sheet_baseline(self, column, row_start):
        self.rtquic1.set_column_list(column, row_start)
        if len(self.rtquic1.get_column_list()) > 0:
            self.rtquic1.setBaseMean()
            self.rtquic1.setBaseSTDEV()
            self.rtquic1.setBaseline()
    def set_result(self, column_start, row):##read row left to right from start point = column_start
        self.rtquic1.set_row_label(row)
        self.rtquic1.set_row_list(column_start, row)
        #check a row_list was generated
        if len(self.rtquic1.get_row_list()) == 0:
            raise ValueError   
        self.rtquic1.set_row_max()
        self.rtquic1.set_time_to_max()
        self.rtquic1.setThreshold()
        self.rtquic1.setLag()
        if len(self.rtquic1.get_column_list()) > 0:
            self.rtquic1.set_time_to_baseline()
            self.rtquic1.set_time_baseline_to_max()
            self.rtquic1.set_gradient()      
    def get_result(self):
        return  (self.rtquic1.get_row_label(),
                self.rtquic1.get_row_max(),
                self.rtquic1.get_time_to_max(),
                self.rtquic1.getLag(),
                self.rtquic1.get_gradient(),
                self.rtquic1.is_positive())    
    def row_filler(self):
        self.row_num += 1
        results = self.get_result()
        print("Results are ",results)
        try:
            for i in range(len(results)):
                self.ws.cell(row = self.row_num,
                        column = i + 1,
                        value = results[i])
        except:
            print("Could not fill a row")
    def data_label_row_filler(self):
        data_labels = ["Label",
                       "Max Val",
                       "Time to Max",
                       "Lag time",
                       "Gradient",
                       "Positive"] 
        self.ws.cell(row =1, column =1, value = self.raw_data)
        try:
            for i in range(len(data_labels)):
                self.ws.cell(row = 2,
                        column = i + 1,
                        value = data_labels[i])
        except:
            print("Could not fill title row")
    def save_wb(self):
        try:
            self.wb.save(self.__str__()+".xlsx")
        except:
            print("An error occurred when I tried to save the file")
            raise FileNotSavedError
    def __str__(self):
        source_file = self.raw_data.replace("-", "_")
        source_file = source_file.replace(" ", "_")
        return ("Analysis_of_File"+ str(self.analysis_num))
    def analyse(self, first_row, last_row, column_start): #first and last row in source file
        self.data_label_row_filler()
        self.set_sheet_baseline(column_start+1, first_row)
        for row in range(first_row, last_row + 1,):
            try:
                self.set_result(column_start, row)
            except ValueError:
                self.save_wb()
                print("had to stop, but file " +self.raw_data + "still saved")
            self.row_filler()
        self.save_wb()

"""Analyses data in an excel file duplicate row by duplicate row
Puts results in a destination excel file"""
class SheetAnalyserMeans(SheetAnalyser):                
    def set_result_list(self, row_start, row):
        self.result_list = []
        self.row_num += 1
        self.set_result(row_start, row)
        Label = self.rtquic1.get_row_label()
        upper_row  = self.get_result()
        self.set_result(row_start, row+1)
        lower_row = self.get_result()
        mean_row = ()
        for i in range (1,len(upper_row)):
            mean_row = mean_row +(((upper_row[i]+ lower_row[i])/2),)   
        maxVal = mean_row[0]
        maxHours = self.rtquic1.hours(mean_row[1])
        lagHours = self.rtquic1.hours(mean_row[2])
        gradient = mean_row[3]
        result = ""
        if upper_row[5] and lower_row[5]:
            result = "Positive"
        elif upper_row[5] or lower_row[5]:
            result = "Negative - interesting finding"
        else:
            result = "Negative"
        #list of mean values with times in hours
        self.result_list = [Label,
                            maxVal,
                            maxHours,
                            lagHours,
                            gradient,
                            result]
    def analyse(self, first_row, last_row, column_start): #first and last row in source file
        self.data_label_row_filler()
        self.set_sheet_baseline(column_start+1, first_row)
        for row in range(first_row, last_row + 1, 2):
            try:
                self.set_result_list(column_start, row)
            except ValueError:
                self.save_wb()
                print("had to stop, but file still saved")
                return
            self.row_filler()
        self.save_wb()
        
"""Analyses data in an excel file sheet row by row
and gets lag time, gradient, max val, time to max
all as floats"""
class SheetAnalyserForCluster(SheetAnalyser):
    def set_result_list(self, column_start, row):
        print("row number sent to set result list "+str(row))
        self.result_list = []
        self.row_num += 1
        self.set_result(column_start, row)
        Label = self.rtquic1.get_row_label()
        maxVal = self.rtquic1.get_row_max()
        maxHours = self.rtquic1.get_time_to_max()
        lagHours = self.rtquic1.getLag()
        gradient = self.rtquic1.get_gradient()
        result = ""
        self.result_list = [Label,
                            maxVal,
                            maxHours,
                            lagHours,
                            gradient]




##METHOD FOR ANALYSING AN INDIVIDUAL EXCEL FILE
row_origin = 13
row_end = 108
column_origin = 4

##def singleSheetAnalysis(RTQUICname, row_origin, row_end, column_origin, means = True):
##    if means == True:
##        method = SheetAnalyserMeans
##    else:
##        method = SheetAnalyser
##    t = method(PATH_TO_FILES+"\\"+"RTQUIC_READ_"+RTQUICname+".xlsx", "All Cycles")                
##    t.analyse(row_origin, row_end, column_origin)
##
##singleSheetAnalysis("19_007", row_origin, row_end, column_origin, means = False)


##METHOD FOR ANALYSING A FOLDER OF EXCEL FILES
t= None
for i in range(len(list_of_files)): 
    if t != None:
        del t 
    t = SheetAnalyser(PATH_TO_FILES+"\\"+list_of_files[i], "Results")
    t.set_analysis_num(i) #keeps track of excel sheets analysed, for generating separate destination files
    t.analyse(13, 108, 14)
    
##ANALYSIS OF ONE COLUMN FROM ONE SHEET

##t = RTQuicSheet(PATH_TO_FILES+"\\"+list_of_files[0], "Results")
##t.set_column_list(14,13)
##t.setBaseMean()
##t.setBaseSTDEV()
##t.setBaseline()
##print("BaseMean is " + str(t.getBaseMean()))
##print("BaseSTDEV is " + str(t.getBaseSTDEV()))
##print("Baseline is " + str(t.getBaseline()))
##t.set_row_list(14,16)
##t.set_row_max()
##print("Row max is " + str(t.get_row_max()))
##t.set_time_to_max()
##print("Time to max is " + str(t.get_time_to_max()))
##t.setThreshold()
##print("Threshold is " + str(t.getThreshold()))
##t.setLag()
##print("Lag is " + str(t.getLag()))
##t.set_time_to_baseline()
##print("Time to baseline is " + str(t.get_time_to_baseline()))
##t.set_time_baseline_to_max()
##print("Time of baseline to max is " + str(t.get_time_baseline_to_max()))
##t.set_gradient()
##print("Gradient is " + str(t.get_gradient()))
